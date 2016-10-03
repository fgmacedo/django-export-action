# coding: utf-8

from __future__ import unicode_literals, absolute_import

from collections import namedtuple
from itertools import chain
import csv
import datetime
import re

from django.db.models import Avg, Count, Sum, Max, Min
from django.http import HttpResponse
from django.utils.text import force_text
from django.template.loader import render_to_string
from django.utils import timezone

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from six import BytesIO, text_type

from .introspection import get_model_from_path_string


DisplayField = namedtuple(
    "DisplayField",
    "path field field_verbose aggregate total group choices field_type",
)


def generate_filename(title, ends_with):
    title = title.split('.')[0]
    title.replace(' ', '_')
    title += ('_' + timezone.now().strftime("%Y-%m-%d_%H%M"))
    if not title.endswith(ends_with):
        title += ends_with
    return title


def add_aggregates(queryset, display_fields):
    agg_funcs = {
        'Avg': Avg, 'Min': Min, 'Max': Max, 'Count': Count, 'Sum': Sum
    }

    for display_field in display_fields:
        if display_field.aggregate:
            func = agg_funcs[display_field.aggregate]
            full_name = display_field.path + display_field.field
            queryset = queryset.annotate(func(full_name))

    return queryset


def report_to_list(queryset, display_fields, user, property_filters=[], preview=False):
    """ Create list from a report with all data filtering.

    queryset: initial queryset to generate results
    display_fields: list of field references or DisplayField models
    user: requesting user
    property_filters: ???
    preview: return only first 50 rows

    Returns list, message in case of issues.
    """
    model_class = queryset.model

    def can_change_or_view(model):
        """ Return True iff `user` has either change or view permission
        for `model`.
        """
        try:
            model_name = model._meta.model_name
        except AttributeError:
            # Needed for Django 1.4.* (LTS).
            model_name = model._meta.module_name

        app_label = model._meta.app_label
        can_change = user.has_perm(app_label + '.change_' + model_name)
        can_view = user.has_perm(app_label + '.view_' + model_name)

        return can_change or can_view

    if not can_change_or_view(model_class):
        return [], 'Permission Denied'

    if isinstance(display_fields, list):
        # Convert list of strings to DisplayField objects.

        new_display_fields = []

        for display_field in display_fields:
            field_list = display_field.split('__')
            field = field_list[-1]
            path = '__'.join(field_list[:-1])

            if path:
                path += '__'  # Legacy format to append a __ here.

            new_model = get_model_from_path_string(model_class, path)
            model_field = new_model._meta.get_field(field)
            choices = model_field.choices
            new_display_fields.append(DisplayField(
                path, field, '', '', None, None, choices, ''
            ))

        display_fields = new_display_fields

    # Build group-by field list.

    group = [df.path + df.field for df in display_fields if df.group]

    # To support group-by with multiple fields, we turn all the other
    # fields into aggregations. The default aggregation is `Max`.

    if group:
        for field in display_fields:
            if (not field.group) and (not field.aggregate):
                field.aggregate = 'Max'

    message = ""
    objects = add_aggregates(queryset, display_fields)

    # Display Values

    display_field_paths = []
    property_list = {}
    display_totals = {}

    for i, display_field in enumerate(display_fields):
        model = get_model_from_path_string(model_class, display_field.path)

        if display_field.field_type == "Invalid":
            continue

        if not model or can_change_or_view(model):
            display_field_key = display_field.path + display_field.field

            if display_field.field_type == "Property":
                property_list[i] = display_field_key
            elif display_field.aggregate == "Avg":
                display_field_key += '__avg'
            elif display_field.aggregate == "Max":
                display_field_key += '__max'
            elif display_field.aggregate == "Min":
                display_field_key += '__min'
            elif display_field.aggregate == "Count":
                display_field_key += '__count'
            elif display_field.aggregate == "Sum":
                display_field_key += '__sum'

            if display_field.field_type not in ('Property', 'Custom Field'):
                display_field_paths.append(display_field_key)

            if display_field.total:
                display_totals[display_field_key] = Decimal(0)

        else:
            message += 'Error: Permission denied on access to {0}.'.format(
                display_field.name
            )

    def increment_total(display_field_key, val):
        """ Increment display total by `val` if given `display_field_key` in
        `display_totals`.
        """
        if display_field_key in display_totals:
            if isinstance(val, bool):
                # True: 1, False: 0
                display_totals[display_field_key] += Decimal(val)
            elif isinstance(val, Number):
                display_totals[display_field_key] += Decimal(str(val))
            elif val:
                display_totals[display_field_key] += Decimal(1)

    # Select pk for primary and m2m relations in order to retrieve objects
    # for adding properties to report rows. Group-by queries do not support
    # Property nor Custom Field filters.

    if not group:
        display_field_paths.insert(0, 'pk')

        m2m_relations = []
        for position, property_path in property_list.items():
            property_root = property_path.split('__')[0]
            root_class = model_class

            try:
                property_root_class = getattr(root_class, property_root)
            except AttributeError:  # django-hstore schema compatibility
                continue

            if type(property_root_class) == ManyToManyDescriptor:
                display_field_paths.insert(1, '%s__pk' % property_root)
                m2m_relations.append(property_root)

    if group:
        values = objects.values(*group)
        values = self.add_aggregates(values, display_fields)
        filtered_report_rows = [
            [row[field] for field in display_field_paths]
            for row in values
        ]
        for row in filtered_report_rows:
            for pos, field in enumerate(display_field_paths):
                increment_total(field, row[pos])
    else:
        filtered_report_rows = []
        values_and_properties_list = []

        values_list = objects.values_list(*display_field_paths)

        for row in values_list:
            row = list(row)
            values_and_properties_list.append(row[1:])
            obj = None  # we will get this only if needed for more complex processing
            # related_objects
            remove_row = False
            # filter properties (remove rows with excluded properties)
            for property_filter in property_filters:
                if not obj:
                    obj = model_class.objects.get(pk=row.pop(0))
                root_relation = property_filter.path.split('__')[0]
                if root_relation in m2m_relations:
                    pk = row[0]
                    if pk is not None:
                        # a related object exists
                        m2m_obj = getattr(obj, root_relation).get(pk=pk)
                        val = reduce(getattr, [property_filter.field], m2m_obj)
                    else:
                        val = None
                else:
                    if property_filter.field_type == 'Custom Field':
                        for relation in property_filter.path.split('__'):
                            if hasattr(obj, root_relation):
                                obj = getattr(obj, root_relation)
                        val = obj.get_custom_value(property_filter.field)
                    else:
                        val = reduce(
                            getattr,
                            (property_filter.path + property_filter.field).split('__'),
                            obj
                        )
                if property_filter.filter_property(val):
                    remove_row = True
                    values_and_properties_list.pop()
                    break
            if not remove_row:
                for i, field in enumerate(display_field_paths[1:]):
                    increment_total(field, row[i + 1])

                for position, display_property in property_list.items():
                    if not obj:
                        obj = model_class.objects.get(pk=row.pop(0))
                    relations = display_property.split('__')
                    root_relation = relations[0]
                    if root_relation in m2m_relations:
                        pk = row.pop(0)
                        if pk is not None:
                            # a related object exists
                            m2m_obj = getattr(obj, root_relation).get(pk=pk)
                            val = reduce(getattr, relations[1:], m2m_obj)
                        else:
                            val = None
                    else:
                        # Could error if a related field doesn't exist
                        try:
                            val = reduce(getattr, relations, obj)
                        except AttributeError:
                            val = None
                    values_and_properties_list[-1].insert(position, val)
                    increment_total(display_property, val)

                filtered_report_rows.append(values_and_properties_list[-1])

            if preview and len(filtered_report_rows) == 50:
                break

    # Sort results if requested.

    if hasattr(display_fields, 'filter'):
        defaults = {
            None: text_type,
            datetime.date: lambda: datetime.date(datetime.MINYEAR, 1, 1),
            datetime.datetime: lambda: datetime.datetime(datetime.MINYEAR, 1, 1),
        }

        # Order sort fields in reverse order so that ascending, descending
        # sort orders work together (based on Python's stable sort). See
        # http://stackoverflow.com/questions/6666748/ for details.

        sort_fields = display_fields.filter(sort__gt=0).order_by('-sort')
        sort_values = sort_fields.values_list('position', 'sort_reverse')

        for pos, reverse in sort_values:
            column = (row[pos] for row in filtered_report_rows)
            type_col = (type(val) for val in column if val is not None)
            field_type = next(type_col, None)
            default = defaults.get(field_type, field_type)()

            filtered_report_rows = sorted(
                filtered_report_rows,
                key=lambda row: self.sort_helper(row[pos], default),
                reverse=reverse,
            )

    values_and_properties_list = filtered_report_rows

    # Build mapping from display field position to choices list.

    choice_lists = {}
    for df in display_fields:
        if df.choices and hasattr(df, 'choices_dict'):
            df_choices = df.choices_dict
            # Insert blank and None as valid choices.
            df_choices[''] = ''
            df_choices[None] = ''
            choice_lists[df.position] = df_choices

    # Build mapping from display field position to format.

    display_formats = {}

    for df in display_fields:
        if hasattr(df, 'display_format') and df.display_format:
            display_formats[df.position] = df.display_format

    def formatter(value, style):
        # Convert value to Decimal to apply numeric formats.
        try:
            value = Decimal(value)
        except Exception:
            pass

        try:
            return style.string.format(value)
        except ValueError:
            return value

    # Iterate rows and convert values by choice lists and field formats.

    final_list = []

    for row in values_and_properties_list:
        row = list(row)

        for position, choice_list in choice_lists.items():
            try:
                row[position] = text_type(choice_list[row[position]])
            except Exception:
                row[position] = text_type(row[position])

        for pos, style in display_formats.items():
            row[pos] = formatter(row[pos], style)

        final_list.append(row)

    values_and_properties_list = final_list

    if display_totals:
        display_totals_row = []

        fields_and_properties = list(display_field_paths[0 if group else 1:])

        for position, value in property_list.items():
            fields_and_properties.insert(position, value)

        for field in fields_and_properties:
            display_totals_row.append(display_totals.get(field, ''))

        # Add formatting to display totals.

        for pos, style in display_formats.items():
            display_totals_row[pos] = formatter(display_totals_row[pos], style)

        values_and_properties_list.append(
            ['TOTALS'] + (len(fields_and_properties) - 1) * ['']
        )
        values_and_properties_list.append(display_totals_row)

    return values_and_properties_list, message


def build_sheet(data, ws, sheet_name='report', header=None, widths=None):
    first_row = 1
    column_base = 1

    ws.title = re.sub(r'\W+', '', sheet_name)[:30]
    if header:
        for i, header_cell in enumerate(header):
            cell = ws.cell(row=first_row, column=i+column_base)
            cell.value = header_cell
            cell.font = Font(bold=True)
            if widths:
                ws.column_dimensions[get_column_letter(i+1)].width = widths[i]

    for row in data:
        for i in range(len(row)):
            item = row[i]
            # If item is a regular string
            if isinstance(item, str):
                # Change it to a unicode string
                try:
                    row[i] = text_type(item)
                except UnicodeDecodeError:
                    row[i] = text_type(item.decode('utf-8', 'ignore'))
            elif type(item) is dict:
                row[i] = text_type(item)
        try:
            ws.append(row)
        except ValueError as e:
            ws.append([e.message])
        except:
            ws.append(['Unknown Error'])


def list_to_workbook(data, title='report', header=None, widths=None):
    """ Create just a openpxl workbook from a list of data """
    wb = Workbook()
    title = re.sub(r'\W+', '', title)[:30]

    if isinstance(data, dict):
        i = 0
        for sheet_name, sheet_data in data.items():
            if i > 0:
                wb.create_sheet()
            ws = wb.worksheets[i]
            build_sheet(
                sheet_data, ws, sheet_name=sheet_name, header=header)
            i += 1
    else:
        ws = wb.worksheets[0]
        build_sheet(data, ws, header=header, widths=widths)
    return wb


def build_xlsx_response(wb, title="report"):
    """ Take a workbook and return a xlsx file response """
    title = generate_filename(title, '.xlsx')
    myfile = BytesIO()
    myfile.write(save_virtual_workbook(wb))
    response = HttpResponse(
        myfile.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % title
    response['Content-Length'] = myfile.tell()
    return response


def list_to_xlsx_response(data, title='report', header=None,
                          widths=None):
    """ Make 2D list into a xlsx response for download
    data can be a 2d array or a dict of 2d arrays
    like {'sheet_1': [['A1', 'B1']]}
    """
    wb = list_to_workbook(data, title, header, widths)
    return build_xlsx_response(wb, title=title)


def list_to_csv_response(data, title='report', header=None, widths=None):
    """ Make 2D list into a csv response for download data.
    """
    response = HttpResponse(content_type="text/csv; charset=UTF-8")
    cw = csv.writer(response)
    for row in chain([header] if header else [], data):
        cw.writerow([force_text(s).encode(response.charset) for s in row])
    return response


def list_to_html_response(data, title='', header=None):
    html = render_to_string('export_action/report_html.html', locals())
    return HttpResponse(html)
